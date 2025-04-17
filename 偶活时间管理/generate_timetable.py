import os
import re
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
import sqlite3

# 设置工作目录
data_dir = "Data"
output_file = "时间管理表.xlsx"

# 生成随机颜色
def random_color():
    r = random.randint(200, 255)  # 限制颜色较浅，以便黑色文字更易读
    g = random.randint(200, 255)
    b = random.randint(200, 255)
    return f"{r:02x}{g:02x}{b:02x}"

# 创建一个组名到颜色的映射
group_colors = {}

# 将字符串时间转换为datetime对象
def parse_time(time_str):
    try:
        return datetime.strptime(time_str, "%H:%M")
    except ValueError:
        return None

# 生成时间轴（5分钟间隔）
def generate_timeline():
    print("正在生成时间轴...")
    timeline = []
    current_time = datetime.strptime("00:00", "%H:%M")
    end_time = datetime.strptime("23:59", "%H:%M")
    
    while current_time <= end_time:
        timeline.append(current_time.strftime("%H:%M"))
        current_time += timedelta(minutes=5)
    
    print(f"时间轴生成完成，共 {len(timeline)} 个时间点")
    return timeline

# 解析文本文件
def parse_file(file_path):
    print(f"正在解析文件: {file_path}")
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read().strip().split('\n')
    
    # 去除空行
    content = [line for line in content if line.strip()]
    
    # 提取城市和场地信息（前两行）
    city = "广州"  # 默认城市
    venue = "未知场馆"  # 默认场馆
    
    if len(content) >= 1 and content[0].startswith("城市"):
        city_parts = content[0].split("：", 1)
        if len(city_parts) > 1:
            city = city_parts[1].strip()
        content = content[1:]  # 移除城市行
    
    if len(content) >= 1 and content[0].startswith("场地"):
        venue_parts = content[0].split("：", 1)
        if len(venue_parts) > 1:
            venue = venue_parts[1].strip()
        content = content[1:]  # 移除场地行
    
    print(f"提取到城市: {city}, 场地: {venue}")
    
    # 检查文件格式（移除城市和场地行后）
    if content and '~' in content[0]:  # 第二种格式（时间范围）
        print(f"检测到第二种格式（时间范围）")
        schedule = parse_format2(content)
    else:  # 第一种格式
        print(f"检测到第一种格式（开始时间+团体名）")
        schedule = parse_format1(content)
    
    print(f"解析完成，共 {len(schedule)} 个演出团体")
    return schedule, city, venue

# 解析第一种格式
def parse_format1(content):
    schedule = []
    i = 0
    
    while i < len(content) - 1:
        start_time = content[i]
        group_name = content[i+1]
        
        # 检查是否有下一个时间点
        if i + 2 < len(content) and re.match(r'\d{1,2}:\d{2}', content[i+2]):
            end_time = content[i+2]
        else:
            # 如果没有下一个时间点，使用与开始时间相同的结束时间
            end_time = start_time
        
        # 处理特典会情况（如果出现两次相同的名称）
        if i > 0 and group_name == content[i-1]:
            # 上一个活动和当前活动名称相同，当前时间是结束时间
            for item in reversed(schedule):
                if item['group'] == group_name:
                    item['end_time'] = start_time
                    break
        else:
            schedule.append({
                'group': group_name,
                'start_time': start_time,
                'end_time': end_time
            })
        
        i += 2  # 移动到下一对时间和组名
    
    # 处理最后一项
    if len(content) % 2 == 1:
        last_time = content[-1]
        for item in reversed(schedule):
            if item['end_time'] == item['start_time']:  # 找到没有正确结束时间的项
                item['end_time'] = last_time
                break
    
    return schedule

# 解析第二种格式
def parse_format2(content):
    schedule = []
    i = 0
    
    while i < len(content) - 1:
        time_range = content[i]
        group_name = content[i+1]
        
        # 解析时间范围
        if '~' in time_range:
            start_time, end_time = time_range.split('~')
        else:
            continue  # 跳过格式不正确的行
        
        schedule.append({
            'group': group_name,
            'start_time': start_time,
            'end_time': end_time
        })
        
        i += 2  # 移动到下一对时间范围和组名
    
    return schedule

# 主函数
def main():
    print("开始生成时间管理表...")
    
    # 获取所有文件
    print(f"扫描目录: {data_dir}")
    all_files = os.listdir(data_dir)
    text_files = [f for f in all_files if f.endswith('.txt')]
    print(f"找到 {len(text_files)} 个文本文件")
    
    # 提取所有日期
    date_pattern = r'【(\d{8})】'
    date_events = {}
    
    for file_name in text_files:
        match = re.search(date_pattern, file_name)
        if match:
            date = match.group(1)
            event_name = file_name.replace(f'【{date}】', '').replace('.txt', '')
            
            if date not in date_events:
                date_events[date] = []
            
            file_path = os.path.join(data_dir, file_name)
            schedule, city, venue = parse_file(file_path)
            
            date_events[date].append({
                'event_name': event_name,
                'schedule': schedule,
                'city': city,
                'venue': venue
            })
    
    print(f"处理了 {len(date_events)} 个日期的活动")
    
    # 生成Excel文件
    print("创建Excel工作簿...")
    wb = Workbook()
    
    # 删除默认创建的sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 为每个日期创建一个sheet
    timeline = generate_timeline()
    
    # 设置数据库文件路径
    db_file = 'timetable.db'
    
    # 创建/连接数据库
    print(f"连接数据库: {db_file}")
    with sqlite3.connect(db_file) as conn:
        c = conn.cursor()
        print("创建数据库表 activities (如果不存在)")
        c.execute('''CREATE TABLE IF NOT EXISTS activities
                     (event_name TEXT, group_name TEXT, 
                     start_time DATETIME, end_time DATETIME, city TEXT, venue TEXT)''')
        
        # 更新数据库逻辑
        print("开始向数据库写入活动数据...")
        activity_count = 0
        
        for date, events in sorted(date_events.items()):
            formatted_date = f"{date[:4]}-{date[4:6]}-{date[6:]}"
            print(f"处理日期: {formatted_date}")
            
            for event in events:
                # 先删除已存在的同名活动数据
                c.execute('DELETE FROM activities WHERE event_name = ?', 
                         (event['event_name'],))
                
                # 插入新数据
                for item in event['schedule']:
                    start_datetime = f"{formatted_date} {item['start_time']}"
                    end_datetime = f"{formatted_date} {item['end_time']}"
                    c.execute('''INSERT INTO activities VALUES
                                 (?, ?, ?, ?, ?, ?)''',
                             (event['event_name'], item['group'],
                              start_datetime, end_datetime, 
                              event['city'], event['venue']))
                    activity_count += 1
        
        print(f"成功写入 {activity_count} 条活动记录到数据库")
        
    # 原有的Excel生成逻辑
    print("开始生成Excel表格...")
    for date, events in sorted(date_events.items()):
        # 添加一个新的sheet
        formatted_date = f"{date[:4]}-{date[4:6]}-{date[6:]}"
        ws = wb.create_sheet(title=formatted_date)
        print(f"创建工作表: {formatted_date}")
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10
        
        # 填充时间轴
        for i, time_str in enumerate(timeline):
            ws.cell(row=i+2, column=1, value=time_str)
        
        # 填充事件
        for col_idx, event in enumerate(events, start=2):
            event_name = event['event_name']
            ws.cell(row=1, column=col_idx, value=event_name)
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            for item in event['schedule']:
                group_name = item['group']
                
                # 为每个组分配唯一颜色（如果尚未分配）
                if group_name not in group_colors:
                    group_colors[group_name] = random_color()
                
                fill = PatternFill(start_color=group_colors[group_name], end_color=group_colors[group_name], fill_type="solid")
                
                # 找到开始和结束时间在时间轴中的索引
                try:
                    start_idx = timeline.index(item['start_time'])
                    
                    # 对于第二种格式，结束时间直接取结束时间
                    # 对于第一种格式，结束时间是下一个开始时间
                    if item['end_time'] != item['start_time']:
                        end_idx = timeline.index(item['end_time'])
                        # 修复：确保结束时间的行也被填充（包含结束时间）
                        end_idx += 1
                    else:
                        # 如果开始时间等于结束时间，将其延长至少5分钟
                        end_idx = start_idx + 1
                        
                    # 特殊情况：如果结束时间比开始时间早，可能是跨越午夜
                    if end_idx <= start_idx:
                        end_idx = len(timeline) - 1  # 延长到当天最后一个时间点
                    
                    # 设置每个单元格的值和填充色
                    for row_idx in range(start_idx, end_idx):
                        cell = ws.cell(row=row_idx+2, column=col_idx, value=group_name)
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 设置一个无边框的默认样式
                        no_border = Border(
                            left=Side(style=None),
                            right=Side(style=None),
                            top=Side(style=None),
                            bottom=Side(style=None)
                        )
                        cell.border = no_border
                    
                    # 只在团体块的外围设置边框
                    # 上边框
                    top_cell = ws.cell(row=start_idx+2, column=col_idx)
                    top_cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style=None)
                    )
                    
                    # 下边框 - 修复：使用正确的行索引
                    bottom_cell = ws.cell(row=start_idx+2+(end_idx-start_idx-1), column=col_idx)
                    bottom_cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style=None),
                        bottom=Side(style='thin')
                    )
                    
                    # 中间行的左右边框
                    for row_idx in range(start_idx+1, end_idx-1):
                        middle_cell = ws.cell(row=row_idx+2, column=col_idx)
                        middle_cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style=None),
                            bottom=Side(style=None)
                        )
                    
                    # 如果只有一行，同时设置上下左右边框
                    if start_idx == end_idx - 1:
                        single_cell = ws.cell(row=start_idx+2, column=col_idx)
                        single_cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                except ValueError:
                    # 如果时间不在时间轴中，跳过
                    print(f"警告: 时间 {item['start_time']} 或 {item['end_time']} 不在时间轴中")
    
    # 冻结第一列和第一行
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = 'B2'
    
    # 保存Excel文件
    print(f"保存Excel文件: {output_file}")
    wb.save(output_file)
    print(f"Excel文件已生成: {output_file}")
    print(f"数据库文件已生成: {db_file}")
    print("所有操作已完成!")

if __name__ == "__main__":
    main()